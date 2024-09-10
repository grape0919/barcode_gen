import os
from barcode.ean import EuropeanArticleNumber13
import openpyxl
from barcode.writer import SVGWriter, SIZE, COMMENT, _set_attributes, create_svg_object

DEFAULT_OPTIONS = {
    "module_width":0.378,
    "module_height":6.8,
    "font_size":9.4,
    "font_path":"./OCRBStd.ttf",
    "quiet_zone":3.3,
    "dpi":300,
}

class CSTWriter(SVGWriter):
    def __init__(self):
        super(CSTWriter, self).__init__()

    def _init(self, code):
        # width, height = self.calculate_size(len(code[0]), len(code))
        width, height = 39.3, 9.8
        self._document = create_svg_object(self.with_doctype)
        self._root = self._document.documentElement
        attributes = {
            "width": SIZE.format(width),
            "height": SIZE.format(height),
        }
        _set_attributes(self._root, **attributes)
        if COMMENT:
            self._root.appendChild(self._document.createComment(COMMENT))
        # create group for easier handling in 3rd party software
        # like corel draw, inkscape, ...
        group = self._document.createElement("g")
        attributes = {"id": "barcode_group"}
        _set_attributes(group, **attributes)
        self._group = self._root.appendChild(group)
        # background = self._document.createElement("rect")
        # attributes = {
        #     "width": "100%",
        #     "height": "100%",
        #     "style": f"fill:{self.background}",
        # }
        # _set_attributes(background, **attributes)
        # self._group.appendChild(background)

    def _create_module(self, xpos, ypos, width, color):
        # Background rect has been provided already, so skipping "spaces"
        if color != self.background:
            element = self._document.createElement("rect")
            height = self.module_height
            if len(self._group.childNodes) in [0,1,14,15,28,29]:
                height = height + 2
            attributes = {
                "x": SIZE.format(xpos),
                "y": SIZE.format(ypos-1.0),
                "width": SIZE.format(width),
                "height": SIZE.format(height),
                "style": f"fill:device-cmyk(0.00, 0.00, 0.00, 100.00);",
            }
            _set_attributes(element, **attributes)
            self._group.appendChild(element)
            
    def _create_text(self, xpos, ypos):
        # check option to override self.text with self.human (barcode as
        # human readable data, can be used to print own formats)
        if self.human != "":
            barcodetext = self.human
        else:
            barcodetext = self.text
        ypos = ypos-3.0
        element = self._document.createElement("text")
        temp_text = barcodetext[0]
        attributes = {
            "x": SIZE.format(-0.5),
            "y": SIZE.format(ypos),
            "style": "font-family:'OCR B Std';font-size:9.4pt;",
        }
        _set_attributes(element, **attributes)
        text_element = self._document.createTextNode(temp_text)
        element.appendChild(text_element)
        self._group.appendChild(element)
        
        
        element = self._document.createElement("text")
        temp_text = barcodetext[1:7]
        attributes = {
            "x": SIZE.format(5.0),
            "y": SIZE.format(ypos),
            "style": "font-family:'OCR B Std';font-size:9.4pt;letter-spacing:0.07em;",
        }
        _set_attributes(element, **attributes)
        text_element = self._document.createTextNode(temp_text)
        element.appendChild(text_element)
        self._group.appendChild(element)
        
        
        element = self._document.createElement("text")
        temp_text = barcodetext[7:]
        attributes = {
            "x": SIZE.format(22.2),
            "y": SIZE.format(ypos),
            "style": "font-family:'OCR B Std';font-size:9.4pt;letter-spacing:0.07em;",
        }
        _set_attributes(element, **attributes)
        text_element = self._document.createTextNode(temp_text)
        element.appendChild(text_element)
        self._group.appendChild(element)
            # ypos += pt2mm(self.font_size) + self.text_line_distance

class MergeWriter(CSTWriter):
    barcode_count = None
    def _create_module(self, xpos, ypos, width, color):
        ypos = ypos + (self.barcode_count * 14.6)
        return super()._create_module(xpos, ypos, width, color)
    def _create_text(self, xpos, ypos):
        ypos = ypos + (self.barcode_count * 14.6)
        return super()._create_text(xpos, ypos)


def read_barcode_list(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    number_list = []

    for i in range(0, sheet.max_row):
        for col in sheet.iter_cols(1, sheet.max_column):
            number_list.append(col[i].value)
            
    return number_list

def code2img(code, output_path, count):
    
    my_writer = CSTWriter()
    merge_writer = MergeWriter()
    str_number = str(code)
    output = os.path.join(output_path, str_number)
    
    if str_number == "None":
        return None, None
    
    my_barcode = EuropeanArticleNumber13(str_number, writer=my_writer)#, writer=image_writer)
    my_barcode.save(output ,DEFAULT_OPTIONS)
    del my_barcode
    
    merge_writer.barcode_count = count
    merge_barcode = EuropeanArticleNumber13(str_number, writer=merge_writer)
    merge_barcode.render(DEFAULT_OPTIONS)
    del merge_barcode
    
    return my_writer._group, merge_writer._group.toxml()
    
